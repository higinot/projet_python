from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from rows_registro import rows_registro
import time
from datetime import datetime
from pytz import timezone

# Nome e caminho dos arquivos excel contendo o input e a base de dados
dir_path = os.path.dirname(os.path.realpath(__file__))
excel_files_folder = r'\excel_files'
input_file = r'\planilha_input.xlsx'
db_file = r'\base.xlsx'
output_folder = r'\outputs'
output_file = Workbook()
output_sheet = output_file.active

# Cria arquivo OUTPUT e nomeia as colunas
output_columns = ['Registro', 'Referencia', 'Referencia_Texto', 'Produto', 'Nome Técnico', 'Fabricante Legal', 'Classificação de Risco', 'Vencimento do Registro', 'Preço', 'Referencia']
for i in range(len(output_columns)):
    col = i+1
    output_sheet.cell(row=1, column=col).value = output_columns[i]

# Carrega o arquivo excel do INPUT e da BASE de dados
input_worksheet = load_workbook(dir_path+excel_files_folder+input_file)
db_worksheet = load_workbook(dir_path+excel_files_folder+db_file)

# Lê o nome de todas as guias existentes na planilha de INPUT
input_sheets = []
for sheet in input_worksheet:
    input_sheets.append(str(sheet)[12:-2])

# Por enquanto, pega somente a PRIMEIRA guia da planilha de INPUT e conta o seu número de linhas
ws = input_worksheet[input_sheets[0]]
row_count = ws.max_row

# Carrega a primeira guia ativa da planilha de BASE e conta o número de linhas
db_ws = db_worksheet.active
db_row_count = db_ws.max_row

# Lista com todos os REGISTROS ANVISA e REFs da planilha de BASE
db_registros = []
db_refs = []
for row in range(1, db_row_count):
    db_registros.append(db_ws['A'+str(row)].value)
    db_refs.append(db_ws['B'+str(row)].value)

# Colunas com Registro Anvisa e Ref na planilha de BASE
anvisa_column = 'J'
ref_column = 'K'

output_row = 2

# Verifica se os Registros ANVISA da planilha de INPUT existem na planilha de BASE
found = [] # Relação das linhas em que foram encontrados os registros
for row in range(2, row_count):
    registro = ws[anvisa_column+str(row)].value
    ref = str(ws[ref_column+str(row)].value)
    
    if registro in db_registros:
        print(f'\nO registro da linha {row} da planilha de INPUT foi encontrado!')
        db_rows_registro = rows_registro(db_registros, registro)

        for db_row in db_rows_registro:
            if str(db_ws['B'+str(db_row)].value) in ref: #Verifica se alguma das refs da planilha de BASE bate com a ref da planilha INPUT
                for col in range(1, 9):
                    output_sheet.cell(row=output_row, column=col).value = str(db_ws[get_column_letter(col)+str(db_row)].value) # Se a ref der match, os dados da planilha de BASE são inseridos na planilha de OUTPUT
                output_sheet.cell(row=output_row, column=9).value = ws["L"+str(row)].value # Adiciona o preço da planilha INPUT na planilha OUTPUT
                output_sheet.cell(row=output_row, column=10).value = ws["AB"+str(row)].value # Adiciona a referencia (TNUMM) da planilha INPUT na planilha OUTPUT
                output_row += 1
                break

        found.append(row)
    else:
        continue
if len(found) == 0:
    print('Nenhum Registro Anvisa do Input foi encontrado na Base!')
else:
    print(f'Foram encontrados os registros das linhas {found}')

zone = 'Brazil/East'
today = datetime.now(timezone(zone))    
d1 = today.strftime("%Y%m%d") #Data, no formato YYYYMMDD
d2 = today.strftime("%H%M%S") #Hora, no formato HHMMSS

output_file.save(dir_path+output_folder+f'/output_{d1}_{d2}.xlsx')
output_file.close()