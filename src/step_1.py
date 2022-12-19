from openpyxl import Workbook, load_workbook, utils
import pyperclip

# Base de dados para higienização
base = load_workbook('base.xlsx')

# Base de dados para teste
teste = load_workbook('teste.xlsx')

# Aba base
planilha = base['Base']

# Selecionando a aba 'Plan1' e a coluna 'A' onde ficam os Anvisas
anvisas = base['Base']['A']

# Selecionando a aba 'Plan1' e a coluna 'A' onde ficam as Referencias
referencias = base['Base']['B']

# Selecionando a aba 'Plan1' e a coluna 'J' onde ficam as Anvisas na planilha teste
anvisas_teste = teste['teste']['J']


# Cria Novo workbook
output = Workbook()

# Seleciona a active Sheet
aba_output = output.active

count=0
# Itera sobre as células da coluna e adiciona os valores à lista
for teste in anvisas_teste:
    for anvisa in anvisas:
        if str(anvisa.value) in str(teste.value):
            count+=1
            row = planilha[anvisa.row]
            values = [cell.value for cell in row]
            for i, value in enumerate(values, start=count):
                output.cell(row=i, column=1, value=value)


output.save('teste_1.xlsx')


